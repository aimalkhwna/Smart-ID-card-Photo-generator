import cv2
import time
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

# ---------------- CONFIG ---------------- #

EXCEL_FILE = "attendance.xlsx"

# ---------------- LOAD FACE DETECTOR ---------------- #
face_cascade = cv2.CascadeClassifier(
    cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
)
if face_cascade.empty():
    print("❌ Haar Cascade not loaded")
    exit()

# ---------------- EXCEL SETUP ---------------- #
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Name", "Date", "Time"])
    wb.save(EXCEL_FILE)

wb = load_workbook(EXCEL_FILE)
ws = wb.active

# ---------------- ATTENDANCE MEMORY ---------------- #
marked_today = set()
today_date = datetime.now().strftime("%Y-%m-%d")

# ---------------- VIDEO ---------------- #
cap = cv2.VideoCapture(0)
if not cap.isOpened():
    print("❌ Camera not accessible")
    exit()

print("📸 Smart Attendance System Started")
print("Press 'Q' to quit")

prev_time = time.time()
name_entered = False
person_name = ""

while True:
    ret, frame = cap.read()
    if not ret:
        break

    frame = cv2.resize(frame, (800, 500))
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    faces = face_cascade.detectMultiScale(
        gray,
        scaleFactor=1.3,
        minNeighbors=5,
        minSize=(80, 80)
    )

    # Ask for name only once
    if len(faces) > 0 and not name_entered:
        person_name = input("Enter Name for Attendance: ").strip()
        name_entered = True

    for (x, y, w, h) in faces:
        cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)

        # Mark attendance once per day
        if name_entered and person_name not in marked_today:
            time_now = datetime.now().strftime("%H:%M:%S")
            ws.append([person_name, today_date, time_now])
            wb.save(EXCEL_FILE)
            marked_today.add(person_name)
            print(f"✅ Attendance Marked for {person_name}")

        cv2.putText(frame, person_name, (x, y - 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)

    # Timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cv2.putText(frame, timestamp, (10, 20),
                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 255), 2)

    # FPS
    curr_time = time.time()
    fps = int(1 / (curr_time - prev_time))
    prev_time = curr_time
    cv2.putText(frame, f"FPS: {fps}", (10, 45),
                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 0), 2)

    cv2.imshow("Smart Attendance System", frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()
wb.close()
print("🛑 Attendance System Stopped")
